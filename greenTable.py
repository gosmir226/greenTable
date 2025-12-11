import sys
import os
import glob
import json
import time
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                             QLineEdit, QTextEdit, QProgressBar, QCheckBox,
                             QMessageBox, QTableWidget, QTableWidgetItem, 
                             QHeaderView, QDialog, QFormLayout, QComboBox)
from PyQt5.QtCore import Qt, QTimer

class TemplateManager:
    def __init__(self, template_file="templates.json"):
        self.template_file = template_file
        self.templates = []
        self.load_templates()
        
    def load_templates(self):
        """Загружает шаблоны из JSON файла"""
        try:
            if os.path.exists(self.template_file):
                with open(self.template_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.templates = data.get('templates', [])
                return True
            else:
                # Создаем пустой файл с шаблонами
                self.save_templates()
                return True
        except Exception as e:
            print(f"Error loading templates: {e}")
            self.templates = []
            return False
    
    def save_templates(self):
        """Сохраняет шаблоны в JSON файл"""
        try:
            with open(self.template_file, 'w', encoding='utf-8') as f:
                json.dump({"templates": self.templates}, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"Error saving templates: {e}")
            return False
    
    def generate_fingerprint(self, cells):
        """Генерирует fingerprint из списка ячеек"""
        # Сортируем ячейки по row, затем col
        sorted_cells = sorted(cells, key=lambda x: (x['row'], x['col']))
        
        # Создаем части fingerprint
        parts = []
        for cell in sorted_cells:
            part = f"{cell['rowspan']}x{cell['colspan']}_{cell['row']}_{cell['col']}_{str(cell['required']).lower()}"
            parts.append(part)
        
        return "|".join(parts)
    
    def find_template(self, sheet_name, group_fingerprint, has_uvnk):
        """Ищет шаблон для группы"""
        for template in self.templates:
            # Проверяем, что подстрока sheet содержится в названии листа
            # И что режим увнк совпадает
            if (template['sheet'] in sheet_name and 
                template['fingerprint'] == group_fingerprint and
                template.get('has_uvnk', False) == has_uvnk):
                return template
        return None
    
    def create_new_template(self, sheet_name, group_cells, has_uvnk, description=""):
        """Создает новый шаблон из группы ячеек"""
        # Создаем cells для fingerprint (без значений)
        cells_for_fingerprint = []
        for cell in group_cells:
            cells_for_fingerprint.append({
                'row': cell['row'],
                'col': cell['col'],
                'rowspan': cell['rowspan'],
                'colspan': cell['colspan'],
                'required': cell['required']
            })
        
        fingerprint = self.generate_fingerprint(cells_for_fingerprint)
        
        # Создаем ячейки для шаблона
        template_cells = []
        for cell in group_cells:
            template_cell = {
                'row': cell['row'],
                'col': cell['col'],
                'rowspan': cell['rowspan'],
                'colspan': cell['colspan'],
                'required': cell['required'],
                'output_column': '',  # Пользователь заполнит позже
                'example': cell.get('value', ''),
                'absolute_position': {
                    'row': cell.get('absolute_row', 0),
                    'col': cell.get('absolute_col', 0)
                }
            }
            template_cells.append(template_cell)
        
        new_template = {
            'id': f'template_{int(time.time())}_{len(self.templates)}',
            'name': f'Автошаблон для {sheet_name}',
            'sheet': sheet_name,
            'has_uvnk': has_uvnk,
            'description': description,
            'fingerprint': fingerprint,
            'cells': template_cells
        }
        
        self.templates.append(new_template)
        self.save_templates()
        
        return new_template
    
    def update_template(self, template_id, updates):
        """Обновляет существующий шаблон"""
        for template in self.templates:
            if template['id'] == template_id:
                template.update(updates)
                self.save_templates()
                return True
        return False

class TemplateEditorDialog(QDialog):
    """Диалог для редактирования шаблона"""
    def __init__(self, template, parent=None):
        super().__init__(parent)
        self.template = template
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle(f"Редактирование шаблона: {self.template['name']}")
        self.setGeometry(200, 200, 800, 500)
        
        layout = QVBoxLayout()
        
        # Информация о шаблоне
        info_label = QLabel(
            f"Лист: {self.template['sheet']} | "
            f"УВНК: {'Да' if self.template.get('has_uvnk', False) else 'Нет'}\n"
            f"Описание: {self.template.get('description', '')}"
        )
        layout.addWidget(info_label)
        
        # Таблица для редактирования ячеек
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            'Row', 'Col', 'RowSpan', 'ColSpan', 'Required', 
            'Пример', 'Абс. позиция', 'Output Column', 'Примечание'
        ])
        
        # Заполняем таблицу
        cells = self.template['cells']
        self.table.setRowCount(len(cells))
        
        for i, cell in enumerate(cells):
            # Основные параметры (только для чтения)
            self.table.setItem(i, 0, QTableWidgetItem(str(cell['row'])))
            self.table.setItem(i, 1, QTableWidgetItem(str(cell['col'])))
            self.table.setItem(i, 2, QTableWidgetItem(str(cell['rowspan'])))
            self.table.setItem(i, 3, QTableWidgetItem(str(cell['colspan'])))
            self.table.setItem(i, 4, QTableWidgetItem(str(cell['required'])))
            
            # Пример значения
            example = cell.get('example', '')
            if example is None:
                example = ''
            example_str = str(example)
            if len(example_str) > 20:
                example_str = example_str[:20] + "..."
            self.table.setItem(i, 5, QTableWidgetItem(example_str))
            
            # Абсолютная позиция
            abs_pos = cell.get('absolute_position', {})
            pos_str = f"R{abs_pos.get('row', 0)}C{abs_pos.get('col', 0)}"
            self.table.setItem(i, 6, QTableWidgetItem(pos_str))
            
            # Редактируемое поле для output_column
            output_col = cell.get('output_column', '')
            item = QTableWidgetItem(output_col if output_col else '')
            self.table.setItem(i, 7, item)
            
            # Примечание (автоматически генерируется)
            note = ""
            if cell['rowspan'] >= 3 and cell['colspan'] >= 8:
                note = "ГИГАНТСКАЯ ЯЧЕЙКА - пропуск группы"
            elif cell['rowspan'] > 1 or cell['colspan'] > 1:
                note = f"Объединение: {cell['rowspan']}×{cell['colspan']}"
            self.table.setItem(i, 8, QTableWidgetItem(note))
        
        # Настраиваем ширину столбцов
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)  # Output Column шире
        
        layout.addWidget(self.table)
        
        # Кнопки
        button_layout = QHBoxLayout()
        self.save_btn = QPushButton("Сохранить")
        self.cancel_btn = QPushButton("Отмена")
        
        self.save_btn.clicked.connect(self.accept)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def get_updated_cells(self):
        """Возвращает обновленные ячейки из таблицы"""
        cells = []
        for i in range(self.table.rowCount()):
            original_cell = self.template['cells'][i]
            cell = {
                'row': int(self.table.item(i, 0).text()),
                'col': int(self.table.item(i, 1).text()),
                'rowspan': int(self.table.item(i, 2).text()),
                'colspan': int(self.table.item(i, 3).text()),
                'required': self.table.item(i, 4).text().lower() == 'true',
                'example': original_cell.get('example', ''),
                'absolute_position': original_cell.get('absolute_position', {}),
                'output_column': self.table.item(i, 7).text()
            }
            cells.append(cell)
        return cells

class ExcelParserApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.template_manager = TemplateManager()
        self.unprocessed_templates = []  # Шаблоны без output_column
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Excel Parser with Template System')
        self.setGeometry(100, 100, 900, 700)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        
        # Directory selection
        dir_layout = QHBoxLayout()
        self.dir_label = QLabel('Selected directory:')
        self.dir_path = QLineEdit()
        self.browse_btn = QPushButton('Browse')
        self.browse_btn.clicked.connect(self.browse_directory)
        dir_layout.addWidget(self.dir_label)
        dir_layout.addWidget(self.dir_path)
        dir_layout.addWidget(self.browse_btn)
        
        # Output file selection
        output_layout = QHBoxLayout()
        self.output_label = QLabel('Output file:')
        self.output_path = QLineEdit('output.xlsx')
        self.output_browse_btn = QPushButton('Browse Output')
        self.output_browse_btn.clicked.connect(self.browse_output_file)
        output_layout.addWidget(self.output_label)
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(self.output_browse_btn)
        
        # Template file selection
        template_layout = QHBoxLayout()
        self.template_label = QLabel('Template file:')
        self.template_path = QLineEdit('templates.json')
        self.template_browse_btn = QPushButton('Browse Template')
        self.template_browse_btn.clicked.connect(self.browse_template_file)
        template_layout.addWidget(self.template_label)
        template_layout.addWidget(self.template_path)
        template_layout.addWidget(self.template_browse_btn)
        
        # Buttons for template management
        template_buttons_layout = QHBoxLayout()
        self.edit_templates_btn = QPushButton('Edit Templates')
        self.edit_templates_btn.clicked.connect(self.edit_templates)
        self.reload_templates_btn = QPushButton('Reload Templates')
        self.reload_templates_btn.clicked.connect(self.reload_templates)
        template_buttons_layout.addWidget(self.edit_templates_btn)
        template_buttons_layout.addWidget(self.reload_templates_btn)
        
        # Auto-create templates option
        self.auto_create_checkbox = QCheckBox("Auto-create new templates for unknown structures")
        self.auto_create_checkbox.setChecked(True)
        
        # Progress bar
        self.progress = QProgressBar()
        
        # Process button
        self.process_btn = QPushButton('Process Excel Files')
        self.process_btn.clicked.connect(self.process_directory)
        
        # Log output
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        
        # Add all to main layout
        layout.addLayout(dir_layout)
        layout.addLayout(output_layout)
        layout.addLayout(template_layout)
        layout.addLayout(template_buttons_layout)
        layout.addWidget(self.auto_create_checkbox)
        layout.addWidget(self.progress)
        layout.addWidget(self.process_btn)
        layout.addWidget(self.log)
        
        central_widget.setLayout(layout)
        
    def browse_directory(self):
        directory = QFileDialog.getExistingDirectory(self, 'Select Directory')
        if directory:
            self.dir_path.setText(directory)
            
    def browse_output_file(self):
        file_name, _ = QFileDialog.getSaveFileName(self, 'Save Output File', 'output.xlsx', 'Excel Files (*.xlsx)')
        if file_name:
            self.output_path.setText(file_name)
            
    def browse_template_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, 'Open Template File', '', 'JSON Files (*.json)')
        if file_name:
            self.template_path.setText(file_name)
            self.reload_templates()
            
    def reload_templates(self):
        """Перезагружает шаблоны из файла"""
        template_file = self.template_path.text()
        self.template_manager.template_file = template_file
        if self.template_manager.load_templates():
            self.log_message(f"Loaded {len(self.template_manager.templates)} templates from {template_file}")
        else:
            self.log_message(f"Failed to load templates from {template_file}")
            
    def edit_templates(self):
        """Открывает диалог для редактирования шаблонов"""
        if not self.template_manager.templates:
            QMessageBox.information(self, "No Templates", "No templates available to edit.")
            return
            
        # Создаем простой диалог для выбора шаблона
        dialog = QDialog(self)
        dialog.setWindowTitle("Select Template to Edit")
        dialog.setGeometry(300, 300, 400, 300)
        
        layout = QVBoxLayout()
        
        # Выпадающий список для выбора шаблона
        form_layout = QFormLayout()
        template_combo = QComboBox()
        for template in self.template_manager.templates:
            uvkn_text = " (УВНК)" if template.get('has_uvnk', False) else ""
            template_combo.addItem(f"{template['name']}{uvkn_text}", template['id'])
        form_layout.addRow("Template:", template_combo)
        
        layout.addLayout(form_layout)
        
        # Кнопки
        button_layout = QHBoxLayout()
        edit_btn = QPushButton("Edit")
        cancel_btn = QPushButton("Cancel")
        
        def on_edit():
            template_id = template_combo.currentData()
            template = next((t for t in self.template_manager.templates if t['id'] == template_id), None)
            if template:
                editor = TemplateEditorDialog(template, self)
                if editor.exec_() == QDialog.Accepted:
                    updated_cells = editor.get_updated_cells()
                    self.template_manager.update_template(template_id, {'cells': updated_cells})
                    self.log_message(f"Template {template['name']} updated")
            dialog.accept()
        
        edit_btn.clicked.connect(on_edit)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(edit_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        dialog.setLayout(layout)
        dialog.exec_()
        
    def log_message(self, message):
        self.log.append(message)
        QApplication.processEvents()
    
    def analyze_group_structure(self, sheet, merged_ranges, group_start_row, col_start, col_end):
        """Анализирует структуру группы и возвращает список ячеек"""
        cells = []
        processed_cells = set()
        
        # Проверяем, есть ли гигантская ячейка, покрывающая всю группу
        for merged_range in merged_ranges:
            if (merged_range.min_row <= group_start_row and 
                merged_range.max_row >= group_start_row + 2 and
                merged_range.min_col <= col_start and 
                merged_range.max_col >= col_end):
                # Гигантская ячейка покрывает всю группу
                self.log_message(f"    WARNING: Giant cell covering entire group found at R{merged_range.min_row}C{merged_range.min_col}")
                return []  # Возвращаем пустой список - группу пропускаем
        
        for row_offset in range(3):  # Группа всегда 3 строки
            row_abs = group_start_row + row_offset
            col_abs = col_start
            
            while col_abs <= col_end:
                # Пропускаем уже обработанные ячейки
                if (row_abs, col_abs) in processed_cells:
                    col_abs += 1
                    continue
                
                # Проверяем, является ли ячейка частью объединения
                is_merged = False
                rowspan = 1
                colspan = 1
                start_row_abs = row_abs
                start_col_abs = col_abs
                
                for merged_range in merged_ranges:
                    if merged_range.min_row <= row_abs <= merged_range.max_row and \
                       merged_range.min_col <= col_abs <= merged_range.max_col:
                        # Это объединенная ячейка
                        is_merged = True
                        rowspan = merged_range.max_row - merged_range.min_row + 1
                        colspan = merged_range.max_col - merged_range.min_col + 1
                        start_row_abs = merged_range.min_row
                        start_col_abs = merged_range.min_col
                        
                        # Проверяем, является ли эта ячейка верхней левой в объединении
                        if row_abs == merged_range.min_row and col_abs == merged_range.min_col:
                            # Это начало объединения - добавляем ячейку
                            cell_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
                            has_data = cell_value is not None and str(cell_value).strip() != ''
                            
                            cells.append({
                                'row': row_offset,
                                'col': col_abs - col_start,
                                'rowspan': rowspan,
                                'colspan': colspan,
                                'required': has_data,
                                'value': cell_value,
                                'absolute_row': merged_range.min_row,
                                'absolute_col': merged_range.min_col
                            })
                        
                        # Помечаем все ячейки этого объединения как обработанные
                        for r in range(merged_range.min_row, merged_range.max_row + 1):
                            for c in range(merged_range.min_col, merged_range.max_col + 1):
                                processed_cells.add((r, c))
                        
                        # Перескакиваем на следующий столбец после объединения
                        col_abs = merged_range.max_col
                        break
                
                if not is_merged:
                    # Обычная ячейка
                    cell_value = sheet.cell(row_abs, col_abs).value
                    has_data = cell_value is not None and str(cell_value).strip() != ''
                    
                    cells.append({
                        'row': row_offset,
                        'col': col_abs - col_start,
                        'rowspan': 1,
                        'colspan': 1,
                        'required': has_data,
                        'value': cell_value,
                        'absolute_row': row_abs,
                        'absolute_col': col_abs
                    })
                    
                    processed_cells.add((row_abs, col_abs))
                
                col_abs += 1
        
        return cells
    
    def extract_data_with_template(self, group_cells, template):
        """Извлекает данные из группы с использованием шаблона"""
        data = {}
        
        # Извлекаем данные по шаблону
        for cell_def in template['cells']:
            output_column = cell_def.get('output_column', '')
            if output_column and output_column.strip():  # Только если output_column задан и не пустой
                # Находим соответствующую ячейку в группе
                for cell in group_cells:
                    if (cell['row'] == cell_def['row'] and 
                        cell['col'] == cell_def['col'] and
                        cell['rowspan'] == cell_def['rowspan'] and
                        cell['colspan'] == cell_def['colspan']):
                        data[output_column] = cell.get('value', '')
                        break
                else:
                    # Если ячейка не найдена, оставляем пустое значение
                    data[output_column] = ''
        
        return data
    
    def process_sheet(self, workbook, sheet_name):
        """Обработка отдельного листа с новой логикой шаблонов"""
        sheet = workbook[sheet_name]
        self.log_message(f"Processing sheet: {sheet_name}")
        
        # Определяем режим работы на основе названия листа
        has_uvnk = 'увнк' in sheet_name.lower()
        self.log_message(f"Режим работы: {'с УВНК' if has_uvnk else 'без УВНК'}")
        
        merged_ranges = list(sheet.merged_cells.ranges)
        self.log_message(f"Found {len(merged_ranges)} merged cell ranges.")
        
        # Вспомогательная функция для получения значения с учетом объединенных ячеек
        def get_cell_value(row, col):
            for merged_range in merged_ranges:
                if merged_range.min_row <= row <= merged_range.max_row and \
                   merged_range.min_col <= col <= merged_range.max_col:
                    return sheet.cell(merged_range.min_row, merged_range.min_col).value
            return sheet.cell(row, col).value
        
        # Находим пустые строки для разделения на цепочки
        empty_rows = []
        for row_idx in range(1, sheet.max_row + 1):
            is_empty = True
            has_fill = False
            
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = get_cell_value(row_idx, col_idx)
                if cell_value is not None and str(cell_value).strip() != '':
                    is_empty = False
                    break
                
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.fill.start_color.index != '00000000':
                    has_fill = True
                    break
            
            if is_empty and not has_fill:
                empty_rows.append(row_idx)
        
        # Разделение на цепочки (chains)
        chain_ranges = []
        start_row = 1
        
        for empty_row in empty_rows:
            if empty_row > start_row:
                chain_ranges.append((start_row, empty_row - 1))
            start_row = empty_row + 1
        
        if start_row <= sheet.max_row:
            chain_ranges.append((start_row, sheet.max_row))
        
        all_data = []
        new_templates_created = []
        
        for chain_idx, (start_row, end_row) in enumerate(chain_ranges):
            chain_height = end_row - start_row + 1
            self.log_message(f"Processing chain {chain_idx+1}: rows {start_row} to {end_row} (высота: {chain_height})")
            
            # Определяем начальный столбец для блоков в зависимости от режима
            first_block_start = 1 if has_uvnk else 2
            
            # Разбиваем цепочку на блоки по 8 столбцов
            for col_start in range(first_block_start, sheet.max_column + 1, 8):
                col_end = min(col_start + 7, sheet.max_column)
                
                # Проверяем, что в цепочке достаточно строк для групп
                if chain_height < 9:  # Минимум 9 строк (3 сверху + 3 группа + 3 снизу)
                    self.log_message(f"  Цепочка слишком короткая ({chain_height} строк), пропускаем")
                    continue
                
                # Получаем дату из первой строки первого столбца блока
                date_value = get_cell_value(start_row, col_start)
                if not date_value:
                    self.log_message(f"  Не найдена дата в ячейке ({start_row}, {col_start}), пропускаем блок")
                    continue
                
                # Определяем начало и конец данных (пропускаем 3 строки сверху и 3 снизу)
                data_start_row = start_row + 3  # Пропускаем 3 строки сверху
                data_end_row = end_row - 3      # Пропускаем 3 строки снизу
                
                if data_start_row > data_end_row:
                    self.log_message(f"  Нет места для данных (data_start_row={data_start_row} > data_end_row={data_end_row})")
                    continue
                
                # Определяем печь
                furnace_value = None
                if has_uvnk:
                    furnace_value = sheet_name
                else:
                    found_uppf = False
                    
                    # Ищем в первых трех строках блока
                    for r in range(start_row, start_row + 3):
                        for c in range(col_start, col_end + 1):
                            cell_val = get_cell_value(r, c)
                            if cell_val and "УППФ" in str(cell_val):
                                furnace_value = cell_val
                                found_uppf = True
                                break
                        if found_uppf:
                            break
                    
                    if not found_uppf:
                        furnace_value = ""
                
                # Определяем количество групп
                available_rows = data_end_row - data_start_row + 1
                num_groups = available_rows // 3
                
                if num_groups <= 0:
                    self.log_message(f"  No groups available in block (available rows: {available_rows})")
                    continue
                
                self.log_message(f"  Found {num_groups} groups in block")
                
                # Обрабатываем каждую группу в блоке
                for group_idx in range(num_groups):
                    group_start_row = data_start_row + group_idx * 3
                    group_end_row = min(group_start_row + 2, data_end_row)
                    
                    # Проверяем, что группа полная (3 строки)
                    if group_end_row - group_start_row < 2:
                        continue
                    
                    # Анализируем структуру группы
                    group_cells = self.analyze_group_structure(
                        sheet, merged_ranges, group_start_row, col_start, col_end
                    )
                    
                    # Если группа пустая (гигантская ячейка пропущена)
                    if not group_cells:
                        self.log_message(f"    Group {group_idx+1}: skipped (giant cell)")
                        continue
                    
                    # Создаем cells для fingerprint (без значений)
                    cells_for_fingerprint = []
                    for cell in group_cells:
                        cells_for_fingerprint.append({
                            'row': cell['row'],
                            'col': cell['col'],
                            'rowspan': cell['rowspan'],
                            'colspan': cell['colspan'],
                            'required': cell['required']
                        })
                    
                    # Генерируем fingerprint
                    group_fingerprint = self.template_manager.generate_fingerprint(cells_for_fingerprint)
                    
                    # Ищем подходящий шаблон
                    template = self.template_manager.find_template(sheet_name, group_fingerprint, has_uvnk)
                    
                    if template:
                        # Используем существующий шаблон
                        group_data = self.extract_data_with_template(group_cells, template)
                        
                        # Добавляем метаданные
                        group_data['Date'] = date_value
                        group_data['Furnace'] = furnace_value
                        group_data['Group'] = group_idx + 1
                        group_data['Block'] = chain_idx + 1
                        group_data['Sheet'] = sheet_name
                        group_data['Template'] = template['id']
                        
                        all_data.append(group_data)
                        
                        self.log_message(f"    Group {group_idx+1}: used template '{template['name']}'")
                    
                    elif self.auto_create_checkbox.isChecked():
                        # Создаем новый шаблон
                        new_template = self.template_manager.create_new_template(
                            sheet_name, 
                            group_cells,
                            has_uvnk,
                            f"Auto-created from sheet {sheet_name}, chain {chain_idx+1}, block starting col {col_start}"
                        )
                        
                        new_templates_created.append(new_template['id'])
                        
                        self.log_message(f"    Group {group_idx+1}: created new template '{new_template['name']}'")
                        
                        # Если в шаблоне уже есть output_column, можно сразу использовать
                        has_output_columns = any(cell.get('output_column', '') for cell in new_template['cells'])
                        if has_output_columns:
                            group_data = self.extract_data_with_template(group_cells, new_template)
                            group_data['Date'] = date_value
                            group_data['Furnace'] = furnace_value
                            group_data['Group'] = group_idx + 1
                            group_data['Block'] = chain_idx + 1
                            group_data['Sheet'] = sheet_name
                            group_data['Template'] = new_template['id']
                            all_data.append(group_data)
                    
                    else:
                        self.log_message(f"    Group {group_idx+1}: no template found and auto-create disabled")
        
        if new_templates_created:
            self.log_message(f"Created {len(new_templates_created)} new templates")
            # Предлагаем пользователю отредактировать новые шаблоны
            QTimer.singleShot(100, self.prompt_template_edit)
        
        return all_data
    
    def prompt_template_edit(self):
        """Предлагает пользователю отредактировать новые шаблоны"""
        if self.template_manager.templates:
            reply = QMessageBox.question(
                self, 
                "New Templates Created", 
                f"Created {len(self.template_manager.templates)} templates. Would you like to edit them now?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes:
                self.edit_templates()
    
    def process_directory(self):
        try:
            directory = self.dir_path.text()
            output_file = self.output_path.text()
            
            if not directory:
                self.log_message("Please select a directory first.")
                return
            
            # Перезагружаем шаблоны
            self.reload_templates()
                
            self.log_message("Scanning directory for Excel files...")
            excel_files = glob.glob(os.path.join(directory, "**", "*.xlsx"), recursive=True)
            excel_files.extend(glob.glob(os.path.join(directory, "**", "*.xls"), recursive=True))
            
            if not excel_files:
                self.log_message("No Excel files found in the directory.")
                return
                
            self.log_message(f"Found {len(excel_files)} Excel files.")
            
            all_data = []
            total_files = len(excel_files)
            
            for file_idx, input_file in enumerate(excel_files):
                self.progress.setValue(int(file_idx / total_files * 100))
                self.log_message(f"Processing file {file_idx+1}/{total_files}: {os.path.basename(input_file)}")
                
                try:
                    # ВСЕГДА используем data_only=True (игнорируем формулы)
                    workbook = load_workbook(filename=input_file, data_only=True)
                    
                    for sheet_name in workbook.sheetnames:
                        sheet_data = self.process_sheet(workbook, sheet_name)
                        all_data.extend(sheet_data)
                        self.log_message(f"  Extracted {len(sheet_data)} rows from {sheet_name}")
                    
                    workbook.close()
                    
                except Exception as e:
                    self.log_message(f"Error processing file {input_file}: {str(e)}")
                    import traceback
                    self.log_message(traceback.format_exc())
            
            if all_data:
                # Собираем все уникальные столбцы
                all_columns = set()
                for row in all_data:
                    all_columns.update(row.keys())
                
                # Сортируем столбцы: сначала метаданные, затем остальные
                base_columns = ['Date', 'Furnace', 'Group', 'Block', 'Sheet', 'Template']
                other_columns = sorted([col for col in all_columns if col not in base_columns])
                final_columns = base_columns + other_columns
                
                # Создаем DataFrame
                df = pd.DataFrame(all_data)
                
                # Убеждаемся, что все столбцы существуют
                for col in final_columns:
                    if col not in df.columns:
                        df[col] = None
                
                # Упорядочиваем столбцы
                df = df[final_columns]
                
                # Сохраняем в Excel
                df.to_excel(output_file, index=False)
                self.log_message(f"Data successfully saved to {output_file}")
                self.log_message(f"Total rows: {len(all_data)}")
                self.log_message(f"Total columns: {len(final_columns)}")
                self.log_message(f"Total templates in library: {len(self.template_manager.templates)}")
                
            else:
                self.log_message("No data was extracted.")
                
            self.progress.setValue(100)
            
        except Exception as e:
            self.log_message(f"Error: {str(e)}")
            import traceback
            self.log_message(traceback.format_exc())
            self.progress.setValue(0)

def main():
    app = QApplication(sys.argv)
    window = ExcelParserApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()