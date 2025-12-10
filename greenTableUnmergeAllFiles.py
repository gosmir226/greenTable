import sys
import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                             QLineEdit, QTextEdit, QProgressBar, QCheckBox)
from PyQt5.QtCore import Qt

class ExcelParserApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Excel Parser for Non-Standard Tables')
        self.setGeometry(100, 100, 800, 600)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        
        # Directory selection
        dir_layout = QHBoxLayout()
        self.dir_label = QLabel('Selected directory:')
        self.dir_path = QLineEdit()
        self.browse_btn = QPushButton('Browse')
        self.browse_btn.clicked.connect(self.browse_directory)
        dir_layout.addWidget(self.dir_label)
        dir_layout.addWidget(self.dir_path)
        dir_layout.addWidget(self.browse_btn)
        
        # Output file selection
        output_layout = QHBoxLayout()
        self.output_label = QLabel('Output file:')
        self.output_path = QLineEdit('output.xlsx')
        self.output_browse_btn = QPushButton('Browse Output')
        self.output_browse_btn.clicked.connect(self.browse_output_file)
        output_layout.addWidget(self.output_label)
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(self.output_browse_btn)
        
        # Data only option
        self.data_only_checkbox = QCheckBox("Read calculated values only (ignore formulas)")
        self.data_only_checkbox.setChecked(True)
        
        # Progress bar
        self.progress = QProgressBar()
        
        # Process button
        self.process_btn = QPushButton('Process Excel Files')
        self.process_btn.clicked.connect(self.process_directory)
        
        # Log output
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        
        # Add all to main layout
        layout.addLayout(dir_layout)
        layout.addLayout(output_layout)
        layout.addWidget(self.data_only_checkbox)
        layout.addWidget(self.progress)
        layout.addWidget(self.process_btn)
        layout.addWidget(self.log)
        
        central_widget.setLayout(layout)
        
    def browse_directory(self):
        directory = QFileDialog.getExistingDirectory(self, 'Select Directory')
        if directory:
            self.dir_path.setText(directory)
            
    def browse_output_file(self):
        file_name, _ = QFileDialog.getSaveFileName(self, 'Save Output File', 'output.xlsx', 'Excel Files (*.xlsx)')
        if file_name:
            self.output_path.setText(file_name)
            
    def log_message(self, message):
        self.log.append(message)
        QApplication.processEvents()
    
    def analyze_first_group_structure(self, sheet, merged_ranges, start_row, col_start, col_end):
        """Анализирует структуру первой группы блока для определения столбцов итоговой таблицы"""
        template_columns = []  # Будет содержать (row_offset, col) для каждого столбца
        processed_cells = set()  # Для отслеживания уже обработанных ячеек
        
        # Проходим по всем ячейкам первых 3 строк группы
        for row_offset in range(3):  # 0, 1, 2 для строк внутри группы
            row = start_row + row_offset
            col = col_start
            
            while col <= col_end:
                # Пропускаем уже обработанные ячейки
                if (row, col) in processed_cells:
                    col += 1
                    continue
                
                # Проверяем, является ли ячейка частью объединения
                is_merged = False
                for merged_range in merged_ranges:
                    if merged_range.min_row <= row <= merged_range.max_row and \
                       merged_range.min_col <= col <= merged_range.max_col:
                        # Это объединенная ячейка
                        is_merged = True
                        # Добавляем только верхнюю левую ячейку объединения
                        if row == merged_range.min_row and col == merged_range.min_col:
                            template_columns.append((row_offset, col))
                        
                        # Помечаем все ячейки этого объединения как обработанные
                        for r in range(merged_range.min_row, merged_range.max_row + 1):
                            for c in range(merged_range.min_col, merged_range.max_col + 1):
                                processed_cells.add((r, c))
                        
                        # Перескакиваем на следующий столбец после объединения
                        col = merged_range.max_col
                        break
                
                if not is_merged:
                    # Обычная ячейка - добавляем ее
                    template_columns.append((row_offset, col))
                    processed_cells.add((row, col))
                
                col += 1
        
        return template_columns
    
    def extract_data_by_template(self, sheet, merged_ranges, template_columns, group_start_row):
        """Извлекает данные из группы по шаблону"""
        group_data = {}
        
        for col_idx, (row_offset, col) in enumerate(template_columns):
            row = group_start_row + row_offset
            
            # Получаем значение с учетом объединенных ячеек
            value = None
            for merged_range in merged_ranges:
                if merged_range.min_row <= row <= merged_range.max_row and \
                   merged_range.min_col <= col <= merged_range.max_col:
                    value = sheet.cell(merged_range.min_row, merged_range.min_col).value
                    break
            
            if value is None:
                value = sheet.cell(row, col).value
            
            group_data[f'Value{col_idx + 1}'] = value
        
        return group_data
    
    def process_sheet(self, workbook, sheet_name):
        """Обработка отдельного листа с новой логикой"""
        sheet = workbook[sheet_name]
        self.log_message(f"Processing sheet: {sheet_name}")
        
        # Определяем режим работы на основе названия листа
        has_uvnk = 'увнк' in sheet_name.lower()
        self.log_message(f"Режим работы: {'с УВНК' if has_uvnk else 'без УВНК'}")
        
        merged_ranges = list(sheet.merged_cells.ranges)
        self.log_message(f"Found {len(merged_ranges)} merged cell ranges.")
        
        # Вспомогательная функция для получения значения с учетом объединенных ячеек
        def get_cell_value(row, col):
            for merged_range in merged_ranges:
                if merged_range.min_row <= row <= merged_range.max_row and \
                   merged_range.min_col <= col <= merged_range.max_col:
                    return sheet.cell(merged_range.min_row, merged_range.min_col).value
            return sheet.cell(row, col).value
        
        # Находим пустые строки для разделения на цепочки
        empty_rows = []
        for row_idx in range(1, sheet.max_row + 1):
            is_empty = True
            has_fill = False
            
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = get_cell_value(row_idx, col_idx)
                if cell_value is not None and str(cell_value).strip() != '':
                    is_empty = False
                    break
                
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.fill.start_color.index != '00000000':
                    has_fill = True
                    break
            
            if is_empty and not has_fill:
                empty_rows.append(row_idx)
        
        # Разделение на цепочки (chains)
        chain_ranges = []
        start_row = 1
        
        for empty_row in empty_rows:
            if empty_row > start_row:
                chain_ranges.append((start_row, empty_row - 1))
            start_row = empty_row + 1
        
        if start_row <= sheet.max_row:
            chain_ranges.append((start_row, sheet.max_row))
        
        all_data = []
        
        for chain_idx, (start_row, end_row) in enumerate(chain_ranges):
            chain_height = end_row - start_row + 1
            self.log_message(f"Processing chain {chain_idx+1}: rows {start_row} to {end_row} (высота: {chain_height})")
            
            # Определяем начальный столбец для блоков в зависимости от режима
            first_block_start = 1 if has_uvnk else 2
            
            # Разбиваем цепочку на блоки по 8 столбцов
            for col_start in range(first_block_start, sheet.max_column + 1, 8):
                col_end = min(col_start + 7, sheet.max_column)
                
                # Проверяем, что в цепочке достаточно строк для групп
                if chain_height < 6:  # Минимум 6 строк (дата + хотя бы одна группа из 3 строк)
                    self.log_message(f"  Цепочка слишком короткая ({chain_height} строк), пропускаем")
                    continue
                
                # Получаем дату из первой строки первого столбца блока (КАК В СТАРОЙ ВЕРСИИ)
                date_value = get_cell_value(start_row, col_start)
                if not date_value:
                    self.log_message(f"  Не найдена дата в ячейке ({start_row}, {col_start}), пропускаем блок")
                    continue
                
                # Определяем начало и конец данных (КАК В СТАРОЙ ВЕРСИИ)
                data_start_row = start_row + 3  # Пропускаем 3 строки сверху
                data_end_row = end_row - 3      # Пропускаем 3 строки снизу
                
                if data_start_row > data_end_row:
                    self.log_message(f"  Нет места для данных (data_start_row={data_start_row} > data_end_row={data_end_row})")
                    continue
                
                # Определяем печь (КАК В СТАРОЙ ВЕРСИИ)
                furnace_value = None
                if has_uvnk:
                    furnace_value = sheet_name
                else:
                    found_uppf = False
                    
                    # Ищем в первых трех строках блока
                    for r in range(start_row, start_row + 3):
                        for c in range(col_start, col_end + 1):
                            cell_val = get_cell_value(r, c)
                            if cell_val and "УППФ" in str(cell_val):
                                furnace_value = cell_val
                                found_uppf = True
                                break
                        if found_uppf:
                            break
                    
                    if not found_uppf:
                        # Ищем в данных
                        for r in range(data_start_row, data_end_row + 1):
                            for c in range(col_start, col_end + 1):
                                cell_val = get_cell_value(r, c)
                                if cell_val and "УППФ" in str(cell_val):
                                    furnace_value = cell_val
                                    found_uppf = True
                                    break
                            if found_uppf:
                                break
                    
                    if not found_uppf:
                        furnace_value = ""
                
                # Анализируем структуру первой группы блока
                template_columns = self.analyze_first_group_structure(
                    sheet, merged_ranges, data_start_row, col_start, col_end
                )
                
                if not template_columns:
                    self.log_message(f"  No template columns found for block starting at column {col_start}")
                    continue
                
                self.log_message(f"  Template has {len(template_columns)} columns")
                
                # Определяем количество групп (КАК В СТАРОЙ ВЕРСИИ)
                # Каждая группа занимает 3 строки
                available_rows = data_end_row - data_start_row + 1
                num_groups = available_rows // 3
                
                if num_groups <= 0:
                    self.log_message(f"  No groups available in block (available rows: {available_rows})")
                    continue
                
                self.log_message(f"  Found {num_groups} groups in block")
                
                # Обрабатываем каждую группу в блоке (КАК В СТАРОЙ ВЕРСИИ)
                for group_idx in range(num_groups):
                    group_start_row = data_start_row + group_idx * 3
                    group_end_row = min(group_start_row + 2, data_end_row)
                    
                    # Проверяем, что группа полная (3 строки)
                    if group_end_row - group_start_row < 2:
                        continue
                    
                    # Извлекаем данные по шаблону
                    group_data = self.extract_data_by_template(
                        sheet, merged_ranges, template_columns, group_start_row
                    )
                    
                    # Добавляем метаданные
                    group_data['Date'] = date_value
                    group_data['Furnace'] = furnace_value
                    group_data['Group'] = group_idx + 1  # Нумерация групп внутри блока
                    group_data['Block'] = chain_idx + 1
                    group_data['Sheet'] = sheet_name
                    
                    all_data.append(group_data)
        
        return all_data
        
    def process_directory(self):
        try:
            directory = self.dir_path.text()
            output_file = self.output_path.text()
            
            if not directory:
                self.log_message("Please select a directory first.")
                return
                
            self.log_message("Scanning directory for Excel files...")
            excel_files = glob.glob(os.path.join(directory, "**", "*.xlsx"), recursive=True)
            excel_files.extend(glob.glob(os.path.join(directory, "**", "*.xls"), recursive=True))
            
            if not excel_files:
                self.log_message("No Excel files found in the directory.")
                return
                
            self.log_message(f"Found {len(excel_files)} Excel files.")
            
            all_data = []
            total_files = len(excel_files)
            
            for file_idx, input_file in enumerate(excel_files):
                self.progress.setValue(int(file_idx / total_files * 100))
                self.log_message(f"Processing file {file_idx+1}/{total_files}: {os.path.basename(input_file)}")
                
                try:
                    if self.data_only_checkbox.isChecked():
                        workbook = load_workbook(filename=input_file, data_only=True)
                    else:
                        workbook = load_workbook(filename=input_file)
                    
                    for sheet_name in workbook.sheetnames:
                        sheet_data = self.process_sheet(workbook, sheet_name)
                        all_data.extend(sheet_data)
                        self.log_message(f"  Extracted {len(sheet_data)} rows from {sheet_name}")
                    
                    workbook.close()
                    
                except Exception as e:
                    self.log_message(f"Error processing file {input_file}: {str(e)}")
            
            if all_data:
                df = pd.DataFrame(all_data)
                
                # Упорядочиваем столбцы
                base_columns = ['Date', 'Furnace', 'Group', 'Block', 'Sheet']
                value_columns = [col for col in df.columns if col.startswith('Value')]
                value_columns.sort(key=lambda x: int(x[5:]) if x[5:].isdigit() else 0)
                
                # Создаем финальный порядок столбцов
                final_columns = base_columns + value_columns
                df = df[final_columns]
                
                # Сохраняем в Excel
                df.to_excel(output_file, index=False)
                self.log_message(f"Data successfully saved to {output_file}")
                self.log_message(f"Total rows: {len(all_data)}")
                self.log_message(f"Total columns: {len(final_columns)}")
                
                # Выводим информацию о структуре
                self.log_message(f"Value columns: {len(value_columns)} (Value1 to Value{len(value_columns)})")
                
            else:
                self.log_message("No data was extracted.")
                
            self.progress.setValue(100)
            
        except Exception as e:
            self.log_message(f"Error: {str(e)}")
            self.progress.setValue(0)

def main():
    app = QApplication(sys.argv)
    window = ExcelParserApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()